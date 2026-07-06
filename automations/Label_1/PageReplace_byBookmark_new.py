import fitz  # PyMuPDF
import os
import re
import shutil
import tempfile
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Bookmark Traget base on replace pages

# Regex patterns for parsing bookmarks and filenames
filename_pattern = re.compile(
    r'(?P<line_id>[\w-]+)_(?P<sheet_number>\d{2})_(?P<revision>00[A-E0-9\-RLM]*)_(?P<engg_file>Y3[\w\d]+)(_(?P<ext_number>\d{3}))?'
)

def parse_filename(filename):
    match = filename_pattern.search(filename)
    if not match:
        return None
    return match.groupdict()

def extract_bookmarks(doc):
    toc = doc.get_toc(simple=True)
    return [{'level': level, 'title': title, 'page': page_num - 1} for level, title, page_num in toc]

def filter_old_bookmarks(toc, replaced_pages_set):
    return [entry for entry in toc if entry['page'] not in replaced_pages_set]

def rebuild_toc(toc_entries):
    sorted_entries = sorted(toc_entries, key=lambda e: (e['page'], e['level']))
    return [[entry['level'], entry['title'], entry['page'] + 1] for entry in sorted_entries]

def add_new_bookmarks(toc, replaced_pages_info):
    for page_num, latest_filename, _level in replaced_pages_info:
        toc.append({'level': 2, 'title': latest_filename, 'page': page_num})

# -------------------------
# NEW helpers for Engg ISO
# -------------------------
def _norm_title(s: str) -> str:
    """Normalize bookmark title to compare reliably across 'Engg ISO', 'Engg_ISO', 'Engg-ISO'."""
    if not isinstance(s, str):
        return ""
    s = s.strip().lower()
    s = s.replace("_", " ").replace("-", " ")
    s = re.sub(r"\s+", " ", s)
    return s

def _is_engg_iso_title(s: str) -> bool:
    return _norm_title(s) == _norm_title("Engg ISO")

def _build_page_to_top(old_toc):
    """Map each bookmark page to its current top-level title (Level-1)."""
    page_to_top = {}
    current_top = None
    for e in old_toc:
        if e['level'] == 1:
            current_top = e['title']
        page_to_top[e['page']] = current_top
    return page_to_top

def _rebuild_toc_sorted(entries):
    entries_sorted = sorted(entries, key=lambda e: (e['page'], e['level']))
    return [[e['level'], e['title'], e['page'] + 1] for e in entries_sorted]

def _add_child_bookmarks(target_list, replaced_pages_info):
    for page_num, latest_filename_base, _old_level in replaced_pages_info:
        target_list.append({'level': 2, 'title': latest_filename_base, 'page': page_num})

# -------------------------
# REPLACED set_bookmarks()
# -------------------------
def set_bookmarks(doc, old_toc, replaced_pages_info):
    """
    Rules:
      - If nothing replaced: keep original TOC.
      - If Engg ISO Level-1 exists: keep EXACTLY one such parent (first encountered) and its page anchor.
        Remove its children that point to replaced pages, then add new Level-2 children for those pages.
      - If Engg ISO doesn't exist: create a parent once at the earliest replaced page and add new children.
      - Keep bookmarks outside Engg ISO untouched.
      - Validate 1-based page numbers when applying set_toc.
    """
    if not replaced_pages_info:
        doc.set_toc(_rebuild_toc_sorted(old_toc))
        return

    replaced_pages_set = {p for p, _, _ in replaced_pages_info}

    # Detect existing Engg ISO parent(s)
    engg_iso_parents = [e for e in old_toc if e['level'] == 1 and _is_engg_iso_title(e['title'])]
    engg_iso_exists = len(engg_iso_parents) > 0

    page_to_top = _build_page_to_top(old_toc)

    kept_outside = []
    kept_engg_children = []
    engg_parent_entry = None
    seen_engg_parent = False

    for e in old_toc:
        if e['level'] == 1 and _is_engg_iso_title(e['title']):
            if not seen_engg_parent:
                engg_parent_entry = {'level': 1, 'title': e['title'], 'page': e['page']}
                seen_engg_parent = True
            # Skip additional Engg ISO parents (avoid duplicates)
            continue

        # Decide if this entry is under Engg ISO by its page's top-level
        top = page_to_top.get(e['page'])
        under_engg = _is_engg_iso_title(top) if top else False

        if under_engg:
            # Remove old child if its page was replaced; otherwise keep it
            if e['page'] in replaced_pages_set:
                continue
            kept_engg_children.append(e)
        else:
            kept_outside.append(e)

    combined = []
    combined.extend(kept_outside)

    if engg_iso_exists and engg_parent_entry:
        # Reuse existing Engg ISO parent at original page
        combined.append(engg_parent_entry)
        # Keep non-replaced children
        combined.extend(kept_engg_children)
        # Add new children for replaced pages
        _add_child_bookmarks(combined, replaced_pages_info)
    else:
        # Create Engg ISO once at earliest replaced page
        first_replaced_page = min(p for p, _, _ in replaced_pages_info)
        combined.append({'level': 1, 'title': 'Engg ISO', 'page': first_replaced_page})
        _add_child_bookmarks(combined, replaced_pages_info)

    # Build final TOC and validate page numbers (1..N inclusive)
    toc_list = _rebuild_toc_sorted(combined)
    total_pages = doc.page_count
    valid_toc = []
    for level, title, page_num in toc_list:
        if 1 <= page_num <= total_pages:
            valid_toc.append([level, title, page_num])
        else:
            print(f"[WARNING] Skipping bookmark '{title}' with invalid page {page_num} in doc with {total_pages} pages.")
    doc.set_toc(valid_toc)

def save_replacement_log(log_entries, log_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Page Replacements Log"
    headers = ['TestPackFile', 'Old_PageNo', 'Old_Bookmark_Title', 'Replaced_FileName', 'Replaced_PageNo', 'Timestamp']
    ws.append(headers)
    for entry in log_entries:
        ws.append(entry)
    for col_num, column_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(len(column_title) + 2, 20)
    wb.save(log_path)

def select_folder_dialog(title):
    root = tk.Tk()
    root.withdraw()
    folder = filedialog.askdirectory(title=title)
    root.destroy()
    return folder

def main():
    root = tk.Tk()
    root.withdraw()
    confirm = messagebox.askokcancel(
        "Confirmation",
        "1. Ensure all latest drawing & pages are resized to its Out Print size .\n"
        "2. Each latest drawing file should have a single page."
    )
    root.destroy()
    if not confirm:
        print("Operation cancelled by user.")
        return

    latest_folder = select_folder_dialog("Select folder for Latest Drawings (PDFs)")
    if not latest_folder:
        print("No latest drawings folder selected. Exiting.")
        return
    testpack_folder = select_folder_dialog("Select folder for Old Test Packs (PDFs)")
    if not testpack_folder:
        print("No old test packs folder selected. Exiting.")
        return

    superseded_folder = os.path.join(os.path.dirname(testpack_folder), "Superseded")
    os.makedirs(superseded_folder, exist_ok=True)
    replace_log_folder = os.path.join(superseded_folder, "Replace log")
    os.makedirs(replace_log_folder, exist_ok=True)

    latest_docs = {}
    for root_dir, _, files in os.walk(latest_folder):
        for file in files:
            if file.lower().endswith(".pdf"):
                latest_docs[file] = os.path.join(root_dir, file)

    testpack_files = []
    for root_dir, _, files in os.walk(testpack_folder):
        for file in files:
            if file.lower().endswith(".pdf"):
                testpack_files.append(os.path.join(root_dir, file))

    log_entries = []

    for tp_path in testpack_files:
        tp_file = os.path.basename(tp_path)
        try:
            if os.path.getsize(tp_path) == 0:
                print(f"Skipping empty file {tp_file}")
                continue
        except Exception as e:
            print(f"Error accessing {tp_file}: {e}")
            continue

        try:
            tp_doc = fitz.open(tp_path)
        except Exception as e:
            print(f"Error opening {tp_file}: {e}")
            continue

        old_toc = extract_bookmarks(tp_doc)
        bm_title_map = {bm['page']: bm['title'] for bm in old_toc}

        new_doc = fitz.open()
        num_pages = tp_doc.page_count
        replaced_pages_info = []

        for page_num in range(num_pages):
            old_bm_title = bm_title_map.get(page_num, "").strip()

            if old_bm_title == "":
                # No bookmark, copy page as is
                new_doc.insert_pdf(tp_doc, from_page=page_num, to_page=page_num)
                continue

            candidate = None
            for latest_file, latest_path in latest_docs.items():
                # Compare file base name (without extension) to bookmark title loosely
                latest_base = os.path.splitext(latest_file)[0].strip().upper()
                bm_title_upper = old_bm_title.upper()

                # Check if bookmark title is contained in the latest file name or vice versa
                if bm_title_upper in latest_base or latest_base in bm_title_upper:
                    candidate = (latest_file, latest_path)
                    break

            if candidate:
                with fitz.open(candidate[1]) as latest_doc:
                    if latest_doc.page_count != 1:
                        print(f"[WARN] Latest drawing {candidate[0]} has more than one page, skipping.")
                        new_doc.insert_pdf(tp_doc, from_page=page_num, to_page=page_num)
                    else:
                        new_doc.insert_pdf(latest_doc, from_page=0, to_page=0)
                        # Determine old bookmark level from TOC for preserving
                        old_level = None
                        for bm in old_toc:
                            if bm['page'] == page_num:
                                old_level = bm['level']
                                break
                        replaced_pages_info.append((page_num, os.path.splitext(candidate[0])[0], old_level))

                        log_entries.append((
                            tp_file,
                            page_num + 1,
                            old_bm_title,
                            candidate[0],
                            page_num + 1,
                            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        ))
            else:
                new_doc.insert_pdf(tp_doc, from_page=page_num, to_page=page_num)

        tp_doc.close()

        if replaced_pages_info:
            superseded_subfolder = os.path.join(superseded_folder, os.path.relpath(os.path.dirname(tp_path), testpack_folder))
            os.makedirs(superseded_subfolder, exist_ok=True)
            superseded_path = os.path.join(superseded_subfolder, tp_file)
            if os.path.exists(tp_path):
                shutil.move(tp_path, superseded_path)
                print(f"Moved original {tp_file} to Superseded folder")

            output_path = tp_path  # overwrite original

            try:
                new_doc.save(output_path)
                new_doc.close()
            except Exception as e:
                print(f"Failed saving replaced PDF {tp_file}: {e}")
                new_doc.close()
                continue

            try:
                final_doc = fitz.open(output_path)
                set_bookmarks(final_doc, old_toc, replaced_pages_info)
                temp_fd, temp_path = tempfile.mkstemp(suffix=".pdf")
                os.close(temp_fd)
                final_doc.save(temp_path)
                final_doc.close()
                shutil.move(temp_path, output_path)
                print(f"Bookmarks updated for replaced file: {output_path}")
            except Exception as e:
                print(f"Failed updating bookmarks for {output_path}: {e}")
        else:
            new_doc.close()

    datetime_str = datetime.now().strftime("%d%m%y_%H%M%S")
    log_filename = f"Replacement_Log_{datetime_str}.xlsx"
    log_path = os.path.join(replace_log_folder, log_filename)
    save_replacement_log(log_entries, log_path)
    print(f"Replacement complete. Log saved to {log_path}")

if __name__ == "__main__":
    main()
