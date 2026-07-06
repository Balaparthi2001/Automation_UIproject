import os
import sys
import signal
import threading
import subprocess
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Excel writer for template
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Optional: data reading
import pandas as pd  # used in the worker for reading input

APP_TITLE = "Excel → Robocopy GUI"
REQUIRED_COLS = ['File Names', 'Target Folder', 'Destination Folder']


class RobocopyApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("900x600")
        self.minsize(800, 520)

        # State
        self.excel_path = tk.StringVar()
        self.is_running = False
        self.stop_event = threading.Event()
        self.worker_thread = None
        self.current_proc: subprocess.Popen | None = None
        self.total_rows = 0
        self.completed_rows = 0

        self._build_ui()
        self._set_initial_state()

    # ---------------- UI ----------------
    def _build_ui(self):
        pad = 8

        # Top: File chooser row
        top_frame = ttk.Frame(self)
        top_frame.pack(fill="x", padx=pad, pady=(pad, 4))

        ttk.Label(top_frame, text="Excel File:").pack(side="left")
        self.entry = ttk.Entry(top_frame, textvariable=self.excel_path)
        self.entry.pack(side="left", fill="x", expand=True, padx=(6, 6))

        self.btn_browse = ttk.Button(top_frame, text="Browse...", command=self.on_browse)
        self.btn_browse.pack(side="left", padx=(0, 6))

        # 👉 NEW: Download Template button
        self.btn_template = ttk.Button(
            top_frame, text="Download Template", command=self.on_download_template
        )
        self.btn_template.pack(side="left")

        # Middle: Controls row
        ctl_frame = ttk.Frame(self)
        ctl_frame.pack(fill="x", padx=pad, pady=(4, 4))

        self.btn_run = ttk.Button(ctl_frame, text="Run", command=self.on_run, width=14)
        self.btn_run.pack(side="left")
        self.btn_stop = ttk.Button(ctl_frame, text="Stop", command=self.on_stop, width=14)
        self.btn_stop.pack(side="left", padx=(8, 0))

        # Status + progress
        status_frame = ttk.Frame(self)
        status_frame.pack(fill="x", padx=pad, pady=(4, 0))

        self.lbl_status = ttk.Label(status_frame, text="Status: Idle")
        self.lbl_status.pack(side="left")

        self.progress = ttk.Progressbar(status_frame, mode="determinate")
        self.progress.pack(side="right", fill="x", expand=True, padx=(10, 0))

        # Log area
        log_frame = ttk.LabelFrame(self, text="Logs")
        log_frame.pack(fill="both", expand=True, padx=pad, pady=(8, pad))

        self.txt_log = tk.Text(log_frame, height=18, wrap="none")
        self.txt_log.pack(side="left", fill="both", expand=True)
        self.txt_log.config(state="disabled")

        yscroll = ttk.Scrollbar(log_frame, orient="vertical", command=self.txt_log.yview)
        yscroll.pack(side="right", fill="y")
        self.txt_log.configure(yscrollcommand=yscroll.set)

    def _set_initial_state(self):
        self.btn_run.state(["disabled"])
        self.btn_stop.state(["disabled"])

    # ------------- UI Helpers -------------
    def log(self, msg: str):
        # Schedule on UI thread
        def _append():
            self.txt_log.config(state="normal")
            self.txt_log.insert("end", msg + "\n")
            self.txt_log.see("end")
            self.txt_log.config(state="disabled")
        self.txt_log.after(0, _append)

    def set_status(self, text: str):
        self.lbl_status.after(0, lambda: self.lbl_status.config(text=f"Status: {text}"))

    def set_buttons_running(self, running: bool):
        if running:
            self.btn_run.state(["disabled"])
            self.btn_browse.state(["disabled"])
            self.btn_template.state(["disabled"])
            self.btn_stop.state(["!disabled"])
        else:
            self.btn_stop.state(["disabled"])
            self.btn_browse.state(["!disabled"])
            self.btn_template.state(["!disabled"])
            # Enable Run only if excel selected and valid
            if self._validate_excel_quick(self.excel_path.get()):
                self.btn_run.state(["!disabled"])
            else:
                self.btn_run.state(["disabled"])

    # ------------- NEW: Download Template -------------
    def on_download_template(self):
        """
        Show save dialog → create a clean Excel template with required headers.
        """
        default_name = "Robocopy_Template.xlsx"
        path = filedialog.asksaveasfilename(
            title="Save Excel Template",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook (*.xlsx)", "*.xlsx")],
        )
        if not path:
            return  # user cancelled

        try:
            self._create_excel_template(path)
            self.log(f"[OK] Template created: {path}")
            messagebox.showinfo(APP_TITLE, f"Template saved:\n{path}")
            # Auto-select the newly created file for convenience
            self.excel_path.set(path)
            if self._validate_excel_quick(path):
                self.btn_run.state(["!disabled"])
        except Exception as e:
            self.log(f"[ERROR] Failed to create template: {e}")
            messagebox.showerror(APP_TITLE, f"Failed to create template:\n{e}")

    def _create_excel_template(self, path: str):
        """
        Build a neat Excel template with REQUIRED_COLS, basic formatting,
        a sample note row (commented as plain text), and freeze the header row.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Robocopy_List"

        # Header row
        header_font = Font(bold=True)
        center = Alignment(vertical="center")
        for col_idx, name in enumerate(REQUIRED_COLS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = header_font
            cell.alignment = center
            # Set reasonable width
            width = 32 if "File Names" in name else 48
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Add a sample hint row (optional)
        ws.cell(row=2, column=1, value="*.pdf or sample.txt")        # File Names
        ws.cell(row=2, column=2, value=r"C:\Source\Folder")          # Target Folder
        ws.cell(row=2, column=3, value=r"D:\Destination\Folder")     # Destination Folder

        # Freeze header
        ws.freeze_panes = "A2"

        # Save
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb.save(path)

    # ------------- Events -------------
    def on_browse(self):
        path = filedialog.askopenfilename(
            title="Select Excel file with file copy list",
            filetypes=[
                ("Excel Workbook (*.xlsx)", "*.xlsx"),
                ("Excel 97-2003 (*.xls)", "*.xls"),
                ("All files", "*.*"),
            ]
        )
        if not path:
            return
        self.excel_path.set(path)
        # Validate columns quickly to decide enabling Run
        if self._validate_excel_quick(path):
            self.log(f"[OK] Selected Excel: {path}")
            self.btn_run.state(["!disabled"])
        else:
            self.log(f"[ERROR] Excel must contain columns: {REQUIRED_COLS}")
            messagebox.showerror(APP_TITLE, f"Excel must contain columns:\n{REQUIRED_COLS}")
            self.btn_run.state(["disabled"])

    def on_run(self):
        if self.is_running:
            return
        path = self.excel_path.get().strip()
        if not path:
            messagebox.showwarning(APP_TITLE, "Please select an Excel file first.")
            return
        if not os.path.exists(path):
            messagebox.showerror(APP_TITLE, "Selected Excel file does not exist.")
            return

        # Reset
        self.stop_event.clear()
        self.completed_rows = 0
        self.progress["value"] = 0
        self.progress["maximum"] = 1  # temporary until we know count
        self.set_status("Processing...")
        self.set_buttons_running(True)
        self.is_running = True
        self.log("=== Started processing ===")

        # Start worker thread
        self.worker_thread = threading.Thread(target=self._worker_run, args=(path,), daemon=True)
        self.worker_thread.start()

    def on_stop(self):
        if not self.is_running:
            return
        self.stop_event.set()
        self.set_status("Stopping...")
        self.log("[INFO] Stop requested by user.")
        # Try to stop current robocopy if running
        proc = self.current_proc
        if proc and proc.poll() is None:
            self._terminate_robocopy(proc)

    # ------------- Excel Reader (Robust) -------------
    def _read_excel_robust(self, path: str, nrows: int | None = None) -> pd.DataFrame:
        """
        Read Excel with explicit engine by extension.
        This avoids failures inside single-file EXEs where pandas can't auto-detect engines.
        """
        ext = os.path.splitext(path)[1].lower()
        if ext == ".xlsx":
            # Requires openpyxl
            return pd.read_excel(path, engine="openpyxl", nrows=nrows)
        elif ext == ".xls":
            # Requires xlrd (note: recent xlrd only supports .xls, not .xlsx)
            return pd.read_excel(path, engine="xlrd", nrows=nrows)
        else:
            raise ValueError("Only .xlsx and .xls files are supported.")

    # ------------- Core Worker -------------
    def _worker_run(self, excel_path: str):
        try:
            try:
                df = self._read_excel_robust(excel_path)
            except ImportError as e:
                self.log(f"[ERROR] Missing Excel engine: {e}")
                self.log("        Install/include 'openpyxl' for .xlsx and 'xlrd' for .xls.")
                self._done(ran=False, msg="Excel engine missing")
                return
            except Exception as e:
                self.log(f"[ERROR] Failed to read Excel: {e}")
                self._done(ran=False, msg="Failed to read Excel")
                return

            # Normalize columns (case-insensitive)
            colmap = {c.lower().strip(): c for c in df.columns}
            required_lower = [c.lower() for c in REQUIRED_COLS]
            if not all(c in colmap for c in required_lower):
                self.log(f"[ERROR] Excel must contain columns: {REQUIRED_COLS}")
                self._done(ran=False, msg="Invalid Excel columns")
                return

            fn_col = colmap['file names']
            src_col = colmap['target folder']
            dst_col = colmap['destination folder']

            # Drop rows missing required values
            df = df.dropna(subset=[fn_col, src_col, dst_col], how="any")
            self.total_rows = len(df.index)
            if self.total_rows == 0:
                self.log("[WARN] No valid rows to process.")
                self._done(ran=True, msg="Nothing to do")
                return

            self.progress.after(0, lambda: self.progress.config(maximum=self.total_rows, value=0))

            # Process each row
            for i, row in df.iterrows():
                if self.stop_event.is_set():
                    self.log("[INFO] Stopping before next item.")
                    break

                file_name = str(row[fn_col]).strip()
                src_folder = str(row[src_col]).strip()
                dest_folder = str(row[dst_col]).strip()
                if not file_name or not src_folder or not dest_folder:
                    self.log(f"[SKIP] Row {i+1}: one or more required fields empty.")
                    self._tick_progress()
                    continue

                # Wildcard?
                has_wildcard = any(ch in file_name for ch in ['*', '?'])
                src_path = os.path.join(src_folder, file_name)

                if not has_wildcard and not os.path.exists(src_path):
                    self.log(f"[SKIP] Row {i+1}: source not found: {src_path}")
                    self._tick_progress()
                    continue

                # Ensure destination
                try:
                    os.makedirs(dest_folder, exist_ok=True)
                except Exception as e:
                    self.log(f"[ERROR] Row {i+1}: cannot create destination '{dest_folder}': {e}")
                    self._tick_progress()
                    continue

                # Build robocopy command
                command = [
                    "robocopy",
                    src_folder,
                    dest_folder,
                    file_name,
                    # Logging tweaks (minimal)
                    "/NFL", "/NDL", "/NJH", "/NJS",
                    # Safer/typical options
                    "/R:1", "/W:1", "/XO", "/FFT"
                ]

                self.log(f"[RUN] Row {i+1}/{self.total_rows}: {file_name}")

                # On Windows, create a new process group so we can send CTRL_BREAK
                creationflags = 0
                if os.name == "nt" and hasattr(subprocess, "CREATE_NEW_PROCESS_GROUP"):
                    creationflags = subprocess.CREATE_NEW_PROCESS_GROUP

                # Start process
                try:
                    self.current_proc = subprocess.Popen(
                        command,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.STDOUT,
                        text=True,
                        bufsize=1,
                        universal_newlines=True,
                        creationflags=creationflags
                    )
                except FileNotFoundError:
                    self.log("[ERROR] 'robocopy' not found. This tool requires Windows.")
                    self._done(ran=False, msg="'robocopy' not found")
                    return
                except Exception as e:
                    self.log(f"[ERROR] Failed to start robocopy: {e}")
                    self._tick_progress()
                    continue

                # Stream output
                try:
                    assert self.current_proc.stdout is not None
                    for line in self.current_proc.stdout:
                        line = line.rstrip()
                        if line:
                            self.log("  " + line)
                        if self.stop_event.is_set():
                            self._terminate_robocopy(self.current_proc)
                            break
                except Exception as e:
                    self.log(f"[WARN] Reading output failed: {e}")

                # Wait end
                rc = self.current_proc.poll()
                if rc is None:
                    rc = self.current_proc.wait()

                if self.stop_event.is_set():
                    self.log(f"[STOPPED] Row {i+1}: interrupted.")
                    self._tick_progress()
                    break

                if rc <= 7:
                    self.log(f"[OK] Row {i+1}: Completed (RC={rc}).")
                else:
                    self.log(f"[FAIL] Row {i+1}: Return code {rc}.")
                self._tick_progress()

            # Done
            if self.stop_event.is_set():
                self._done(ran=True, msg="Stopped by user")
            else:
                self._done(ran=True, msg="Completed")

        except Exception as e:
            self.log(f"[FATAL] Unexpected error: {e}")
            self._done(ran=False, msg="Unexpected error")

    def _terminate_robocopy(self, proc: subprocess.Popen):
        try:
            if os.name == "nt":
                try:
                    # Works if created with CREATE_NEW_PROCESS_GROUP
                    proc.send_signal(signal.CTRL_BREAK_EVENT)
                except Exception:
                    pass
            # Give a moment to exit
            try:
                proc.wait(timeout=2)
            except Exception:
                pass
            if proc.poll() is None:
                proc.kill()
        except Exception as e:
            self.log(f"[WARN] Failed to terminate robocopy: {e}")

    def _tick_progress(self):
        self.completed_rows += 1
        val = self.completed_rows
        self.progress.after(0, lambda: self.progress.config(value=val))

    def _done(self, ran: bool, msg: str):
        # Reset running state, buttons, status
        def _finish():
            self.is_running = False
            self.current_proc = None
            self.set_buttons_running(False)
            self.set_status("Idle" if ran else "Error")
            self.log(f"=== {msg} ===")
        self.after(0, _finish)

    # Quick header validation for enabling Run button after Browse
    def _validate_excel_quick(self, path: str) -> bool:
        if not path or not os.path.exists(path):
            return False
        try:
            # Fast read header only
            df = self._read_excel_robust(path, nrows=1)
            colmap = {c.lower().strip(): c for c in df.columns}
            return all(c.lower() in colmap for c in REQUIRED_COLS)
        except Exception:
            return False


def main():
    if os.name != "nt":
        # GUI message; safe even before Tk mainloop
        try:
            messagebox.showerror(APP_TITLE, "This tool requires Windows (robocopy).")
        except Exception:
            print("This tool requires Windows (robocopy).", file=sys.stderr)
        return
    app = RobocopyApp()
    app.mainloop()


if __name__ == "__main__":
    main()